import io
import openpyxl
import pytest
from bot.main import process_excel_file_bytes

@pytest.mark.asyncio
async def test_process_excel_file_bytes():
    # 1. Create a mock excel workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Numbers Sheet"

    # Set up some rows with range and numbers
    ws.cell(row=1, column=1, value="iVAS Premium rate SMS My Numbers")
    ws.cell(row=3, column=1, value="Range")
    ws.cell(row=3, column=2, value="Number")
    ws.cell(row=3, column=3, value="A2P")

    # Egypt sample
    ws.cell(row=4, column=1, value="EGYPT 13450")
    ws.cell(row=4, column=2, value=201551060331)

    # Indonesia sample
    ws.cell(row=5, column=1, value="INDO combo")
    ws.cell(row=5, column=2, value="0812345678")

    # USA sample
    ws.cell(row=6, column=1, value="US combo")
    ws.cell(row=6, column=2, value="12025550123")

    excel_in_fp = io.BytesIO()
    wb.save(excel_in_fp)
    input_bytes = excel_in_fp.getvalue()

    # 2. Invoke our parser
    output_bytes, cleaned_numbers = await process_excel_file_bytes(input_bytes, chat_id=12345)

    # 3. Assert on cleaned numbers
    assert len(cleaned_numbers) == 3
    assert "+201551060331" in cleaned_numbers
    assert "+62812345678" in cleaned_numbers
    assert "+12025550123" in cleaned_numbers

    # 4. Parse the output excel sheet
    out_wb = openpyxl.load_workbook(io.BytesIO(output_bytes))
    out_ws = out_wb.active

    # Headers are: Flag, Range, Number, A2P, WA Status
    assert out_ws.cell(row=1, column=1).value == "Flag"
    assert out_ws.cell(row=1, column=2).value == "Range"
    assert out_ws.cell(row=1, column=3).value == "Number"
    assert out_ws.cell(row=1, column=4).value == "A2P"
    assert out_ws.cell(row=1, column=5).value == "WA Status"

    # Row 2 should be Egypt: 🇪🇬, EGYPT 13450, +201551060331
    assert out_ws.cell(row=2, column=1).value == "🇪🇬"
    assert out_ws.cell(row=2, column=2).value == "EGYPT 13450"
    assert out_ws.cell(row=2, column=3).value == "+201551060331"

    # Row 3 should be Indo: 🇮🇩, INDO combo, +62812345678
    assert out_ws.cell(row=3, column=1).value == "🇮🇩"
    assert out_ws.cell(row=3, column=2).value == "INDO combo"
    assert out_ws.cell(row=3, column=3).value == "+62812345678"

    # Row 4 should be US: 🇺🇸, US combo, +12025550123
    assert out_ws.cell(row=4, column=1).value == "🇺🇸"
    assert out_ws.cell(row=4, column=2).value == "US combo"
    assert out_ws.cell(row=4, column=3).value == "+12025550123"
